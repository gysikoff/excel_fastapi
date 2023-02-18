from fastapi import FastAPI, UploadFile, HTTPException, status, Request
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import uvicorn

from openpyxl import load_workbook
from io import BytesIO
from typing import List


app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="page")


@app.get('/', response_class=HTMLResponse)
def root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.post('/upload/')
def upload_excel(excel: UploadFile, columns: List[str]):
    if not excel.filename.endswith('.xlsx'): # Checks file extension
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail='Wrong file extension'
        )
    
    response = {'file': excel.filename, 'summary': []} # Creates response object

    columns = list(columns[0].split(','))
    wb = load_workbook(filename=BytesIO(excel.file.read())) # Load SpooledTemporaryFile to openpyxl

    cell_value_list = []

    append_list = False
    # Reads values in every column of excel file
    for ws in wb.worksheets:
            for col in ws.iter_cols():
                append_list = False
                for cell in col:
                    if cell.value is not None:
                        if not append_list:
                            if str(cell.value).strip() in columns: # Finds a cell via cell.value provided in request param column
                                col_name = str(cell.value).strip()
                                append_list = True # sets the boolean to true so every float or int from cells below is added to a list
                        else:
                            if type(cell.value) == int or type(cell.value) == float:
                                cell_value_list.append(cell.value)
                if append_list:
                    search = {'column': col_name, 'sum': round(sum(cell_value_list), 2), 'avg': round(avg(cell_value_list), 2)} # creates result object to be outputed in response
                    response['summary'].append(search)
                    cell_value_list = []

    return response


if __name__ == '__main__':
    uvicorn.run(app, host='127.0.0.1', port=8000)


def sum(list):
    sum = 0
    for x in list:
        sum += x

    return sum

def avg(list):
    sum = 0
    i = 0
    for x in list:
        sum +=x
        i += 1

    return sum/i